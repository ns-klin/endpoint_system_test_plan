#!/usr/bin/env python3
"""
Export all test cases to a single .xlsx file with one sheet per chapter.
Format matches the reference: Device Classification Service - System test scenarios.

Columns:
  TC ID | Test Case Name | Feature Under Test | What We're Testing |
  Why It Matters | What Could Go Wrong | Platform | Severity |
  Auto Priority | Gap Type | Related Bugs | Preconditions |
  Test Steps | Expected Result | Failure Indicators | Risk if Untested |
  Automation Coverage
"""

import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from generate_missing_tcs import CH03_TCS, CH05_TCS, CH07_TCS, CH11_TCS

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.normpath(os.path.join(SCRIPT_DIR, '..'))
CHAPTERS_DIR = os.path.join(PROJECT_ROOT, 'chapters')
TC_EXPORT_DIR = os.path.join(PROJECT_ROOT, 'tc_export')
OUTPUT_FILE = os.path.join(TC_EXPORT_DIR, "NSClient_Test_Cases_Full.xlsx")

CHAPTERS = [
    ("01_installation.md", "01 Installation"),
    ("02_enrollment.md", "02 Enrollment"),
    ("03_service_lifecycle.md", "03 Service Lifecycle"),
    ("04_config_download.md", "04 Config Download"),
    ("05_steering_config.md", "05 Steering Config"),
    ("06_client_status.md", "06 Client Status"),
    ("07_tunnel_management.md", "07 Tunnel Mgmt"),
    ("08_gateway_selection.md", "08 Gateway Selection"),
    ("09_traffic_steering.md", "09 Traffic Steering"),
    ("10_bypass.md", "10 Bypass"),
    ("11_failclose.md", "11 FailClose"),
    ("12_device_classification.md", "12 Device Classification"),
    ("13_certificate_management.md", "13 Certificate Mgmt"),
    ("14_proxy_management.md", "14 Proxy Management"),
    ("15_npa_integration.md", "15 NPA Integration"),
    ("16_dem.md", "16 DEM"),
    ("17_ipc_nscom2.md", "17 IPC NSCom2"),
    ("18_security.md", "18 Security"),
    ("19_integration_architecture.md", "19 Integration"),
    ("20_supportability.md", "20 Supportability"),
    ("21_watchdog.md", "21 Watchdog"),
]

# Pre-generated TCs for chapters that lack TC definitions in their .md files
GENERATED_TCS = {
    "03": CH03_TCS,
    "05": CH05_TCS,
    "07": CH07_TCS,
    "11": CH11_TCS,
}

HEADERS = [
    "TC ID",
    "Test Case Name",
    "Feature Under Test",
    "What We're Testing",
    "Why It Matters",
    "What Could Go Wrong",
    "Platform",
    "Severity",
    "Auto Priority",
    "Gap Type",
    "Related Bugs",
    "Preconditions",
    "Test Steps",
    "Expected Result",
    "Failure Indicators",
    "Risk if Untested",
    "Automation Coverage",
]

# Column widths
COL_WIDTHS = [10, 30, 20, 35, 35, 35, 15, 8, 10, 12, 20, 35, 60, 40, 35, 35, 25]


def get_platform_from_context(content, pos):
    """Determine platform based on section headers above position."""
    before = content[:pos]
    patterns = [
        (r'##\s+Windows', 'Windows'),
        (r'##\s+macOS', 'macOS'),
        (r'##\s+Linux', 'Linux'),
        (r'##\s+Android', 'Android'),
        (r'##\s+iOS', 'iOS'),
        (r'##\s+ChromeOS', 'ChromeOS'),
        (r'##\s+Backend', 'Backend'),
        (r'##\s+Cross-Platform', 'Cross-Platform'),
        (r'###\s+Windows Test Cases', 'Windows'),
        (r'###\s+macOS Test Cases', 'macOS'),
        (r'###\s+Linux Test Cases', 'Linux'),
        (r'###\s+Android Test Cases', 'Android'),
        (r'###\s+Cross-Platform Test Cases', 'Cross-Platform'),
        (r'###\s+Cross-Flow Test Cases', 'Cross-Flow'),
    ]
    last_pos = -1
    platform = ""
    for pat, plat in patterns:
        for m in re.finditer(pat, before):
            if m.start() > last_pos:
                last_pos = m.start()
                platform = plat
    return platform


def get_chapter_feature(filename):
    """Get feature name from chapter filename."""
    features = {
        "01": "Installation & Upgrade",
        "02": "Enrollment",
        "03": "Service Lifecycle",
        "04": "Config Download & Management",
        "05": "Steering Config",
        "06": "Client Status Reporting",
        "07": "Tunnel Management",
        "08": "Gateway Selection (GSLB)",
        "09": "Traffic Interception & Steering",
        "10": "Bypass Mechanisms",
        "11": "FailClose",
        "12": "Device Classification",
        "13": "Certificate Management",
        "14": "Proxy Detection & Management",
        "15": "NPA / Private Access",
        "16": "DEM (Digital Experience Monitoring)",
        "17": "IPC (NSCom2/NSMsg2)",
        "18": "Security Mechanisms",
        "19": "Multi-Component Integration",
        "20": "Supportability & Diagnostics",
        "21": "Watchdog",
    }
    return features.get(filename[:2], "Unknown")


def parse_detailed_tc(content, chapter_num):
    """Parse test cases in **TC-XX-XX: Title** format."""
    test_cases = []
    # Match both TC-XX-XX and TC-CF-XX-XX
    tc_pattern = re.compile(r'\*\*(TC-(?:CF-)?\d{2}-\d{2}):\s*(.+?)\*\*')
    matches = list(tc_pattern.finditer(content))

    for i, match in enumerate(matches):
        tc_id = match.group(1)
        tc_title = match.group(2).strip()

        start = match.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else min(start + 3000, len(content))
        block = content[start:end]

        tc = {"id": tc_id, "name": tc_title}

        # Extract fields from | **Field** | Value | table
        def extract_field(field_name):
            pat = re.search(r'\*\*' + field_name + r'\*\*\s*\|\s*(.+?)(?:\s*\|?\s*$)', block, re.MULTILINE)
            return pat.group(1).strip() if pat else ""

        tc["severity"] = extract_field("Severity")
        tc["related_bugs"] = extract_field("Related Bugs")
        tc["flow_point"] = extract_field("Flow Point")
        tc["gap_type"] = extract_field("Gap Type")
        tc["auto_priority"] = extract_field("Automation Priority")

        platform_field = extract_field("Platforms?")
        if not platform_field:
            platform_field = get_platform_from_context(content, match.start())
        tc["platform"] = platform_field

        # Preconditions
        precond_match = re.search(r'\*\*Preconditions\*\*:\s*(.+?)(?=\n\*\*Steps)', block, re.DOTALL)
        tc["preconditions"] = precond_match.group(1).strip() if precond_match else ""

        # Steps - extract numbered list
        steps_match = re.search(r'\*\*Steps\*\*:\s*\n((?:\d+\..+\n?)+)', block)
        if steps_match:
            steps_lines = steps_match.group(1).strip().split('\n')
            tc["steps"] = " | ".join(s.strip() for s in steps_lines)
        else:
            tc["steps"] = ""

        # Expected Result
        expected_match = re.search(r'\*\*Expected Result\*\*:\s*(.+?)(?=\n\*\*|\n---|\n\n)', block, re.DOTALL)
        tc["expected_result"] = expected_match.group(1).strip().replace('\n', ' ') if expected_match else ""

        # Failure Indicators
        fail_match = re.search(r'\*\*Failure Indicators\*\*:\s*(.+?)(?=\n\*\*|\n---|\n\n)', block, re.DOTALL)
        tc["failure_indicators"] = fail_match.group(1).strip().replace('\n', ' ') if fail_match else ""

        # Risk if Untested
        risk_match = re.search(r'\*\*Risk if Untested\*\*:\s*(.+?)(?=\n\*\*|\n---|\n\n)', block, re.DOTALL)
        tc["risk"] = risk_match.group(1).strip().replace('\n', ' ') if risk_match else ""

        # Automation Coverage
        auto_match = re.search(r'\*\*Automation Coverage\*\*:\s*(.+?)(?=\n\*\*|\n---|\n\n)', block, re.DOTALL)
        tc["automation"] = auto_match.group(1).strip().replace('\n', ' ') if auto_match else ""

        test_cases.append(tc)

    return test_cases


def parse_heading_tc(content, chapter_num):
    """Parse test cases in ### TC-XX-XX: Title format (Ch04)."""
    test_cases = []
    tc_pattern = re.compile(r'###\s+(TC-\d{2}-\d{2}):\s*(.+)')
    matches = list(tc_pattern.finditer(content))

    for i, match in enumerate(matches):
        tc_id = match.group(1)
        tc_title = match.group(2).strip()

        start = match.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else min(start + 3000, len(content))
        block = content[start:end]

        tc = {"id": tc_id, "name": ""}

        def extract_field(field_name):
            pat = re.search(r'\*\*' + field_name + r'\*\*\s*\|\s*(.+?)(?:\s*\|?\s*$)', block, re.MULTILINE)
            return pat.group(1).strip() if pat else ""

        tc["name"] = extract_field("Test Case") or tc_title
        tc["severity"] = extract_field("Severity")
        tc["related_bugs"] = extract_field("Related Bugs")
        tc["flow_point"] = extract_field("Flow Point")
        tc["gap_type"] = extract_field("Gap Type")
        tc["auto_priority"] = extract_field("Automation Priority")
        tc["platform"] = get_platform_from_context(content, match.start())

        # Preconditions from table format
        precond = extract_field("Preconditions")
        tc["preconditions"] = precond

        # Steps from table format
        steps = extract_field("Steps")
        tc["steps"] = steps.replace("<br/>", " | ") if steps else ""

        # Expected Result
        expected = extract_field("Expected Result")
        tc["expected_result"] = expected

        # Failure Indicators
        fail = extract_field("Failure Indicators")
        tc["failure_indicators"] = fail.replace("- ", "| ").replace("<br/>", " | ") if fail else ""

        # Risk if Untested
        tc["risk"] = extract_field("Risk if Untested")
        tc["automation"] = ""

        test_cases.append(tc)

    return test_cases


def parse_table_tc(content, chapter_num):
    """Parse test cases in markdown table format."""
    test_cases = []
    lines = content.split('\n')

    for line_idx, line in enumerate(lines):
        tc_match = re.match(r'\|\s*(TC-(?:CF-)?\d{2}-\d{2})\s*\|(.+)', line)
        if not tc_match:
            continue

        tc_id = tc_match.group(1)
        rest = tc_match.group(2)
        cells = [c.strip() for c in rest.split('|') if c.strip()]

        tc = {"id": tc_id}

        if chapter_num == "06":
            tc["name"] = cells[0] if len(cells) > 0 else ""
            tc["severity"] = cells[1] if len(cells) > 1 else ""
            tc["related_bugs"] = cells[2] if len(cells) > 2 else ""
            tc["flow_point"] = cells[3] if len(cells) > 3 else ""
            tc["gap_type"] = cells[4] if len(cells) > 4 else ""
            tc["auto_priority"] = cells[5] if len(cells) > 5 else ""
        elif chapter_num == "13":
            tc["name"] = cells[0] if len(cells) > 0 else ""
            tc["severity"] = cells[1] if len(cells) > 1 else ""
            tc["related_bugs"] = cells[2] if len(cells) > 2 else ""
            tc["flow_point"] = cells[3] if len(cells) > 3 else ""
            tc["gap_type"] = cells[4] if len(cells) > 4 else ""
            tc["auto_priority"] = cells[5] if len(cells) > 5 else ""
        elif chapter_num in ("09", "15", "18"):
            tc["name"] = cells[0] if len(cells) > 0 else ""
            tc["severity"] = cells[1] if len(cells) > 1 else ""
            tc["auto_priority"] = cells[2] if len(cells) > 2 else ""
            tc["gap_type"] = cells[3] if len(cells) > 3 else ""
            platform_cell = cells[4] if len(cells) > 4 else ""
            tc["platform"] = platform_cell
            tc["related_bugs"] = ""
            tc["flow_point"] = ""
        else:
            # Generic: | TC | desc | severity | priority | gap_type |
            tc["name"] = cells[0] if len(cells) > 0 else ""
            tc["severity"] = cells[1] if len(cells) > 1 else ""
            tc["auto_priority"] = cells[2] if len(cells) > 2 else ""
            tc["gap_type"] = cells[3] if len(cells) > 3 else ""
            tc["related_bugs"] = ""
            tc["flow_point"] = ""

        # Fill defaults
        tc.setdefault("platform", "")
        tc.setdefault("related_bugs", "")
        tc.setdefault("flow_point", "")
        tc.setdefault("preconditions", "")
        tc.setdefault("steps", "")
        tc.setdefault("expected_result", "")
        tc.setdefault("failure_indicators", "")
        tc.setdefault("risk", "")
        tc.setdefault("automation", "")

        # Get platform from context if not set
        if not tc["platform"]:
            # Calculate approximate position in content
            pos = content.find(line)
            if pos >= 0:
                tc["platform"] = get_platform_from_context(content, pos)

        test_cases.append(tc)

    return test_cases


def generate_steps_for_table_tc(tc, chapter_feature):
    """Generate reasonable test steps for test cases that only have a table-format title."""
    name = tc.get("name", "")
    platform = tc.get("platform", "")
    flow_point = tc.get("flow_point", "")
    related_bugs = tc.get("related_bugs", "")

    steps = []
    # General framework for all test cases
    steps.append(f"1. SETUP: Configure {chapter_feature} test environment on {platform or 'target platform'}")

    if flow_point:
        steps.append(f"2. TRIGGER: Execute action that invokes {flow_point}")
    else:
        steps.append(f"2. TRIGGER: Execute the scenario: {name}")

    steps.append(f"3. VERIFY: Confirm the expected behavior occurs without errors")
    steps.append(f"4. MONITOR: Check nsdebuglog.log for relevant keywords and error patterns")

    if related_bugs and related_bugs != "--" and "None" not in related_bugs and "(Predicted" not in related_bugs:
        steps.append(f"5. REGRESSION: Verify fix for {related_bugs} is not regressed")
    else:
        steps.append(f"5. NEGATIVE: Test failure scenarios and verify graceful handling")

    steps.append(f"6. VALIDATE: Confirm client returns to normal operation after test")

    return " | ".join(steps)


def generate_expected_result(tc, chapter_feature):
    """Generate expected result for table-format test cases."""
    name = tc.get("name", "")
    return f"{name} completes successfully without errors; client remains stable and functional"


def generate_preconditions(tc, chapter_feature):
    """Generate preconditions for table-format test cases."""
    platform = tc.get("platform", "")
    name = tc.get("name", "").lower()

    parts = []
    if platform:
        parts.append(f"{platform} device with NSClient installed and enrolled")
    else:
        parts.append("Device with NSClient installed and enrolled")

    if "tunnel" in name or "dem" in name:
        parts.append("tunnel connected")
    if "failclose" in name:
        parts.append("FailClose enabled in admin console")
    if "npa" in name:
        parts.append("NPA enabled with private app access configured")
    if "vdi" in name:
        parts.append("VDI/multi-user environment with 2+ active sessions")
    if "upgrade" in name:
        parts.append("upgrade package available on MP")
    if "config" in name:
        parts.append("admin console access for policy changes")

    return "; ".join(parts)


def enrich_tc(tc, chapter_feature):
    """Generate 'What We're Testing', 'Why It Matters', 'What Could Go Wrong' from available data."""
    name = tc.get("name", "")
    flow_point = tc.get("flow_point", "")
    related_bugs = tc.get("related_bugs", "")
    gap_type = tc.get("gap_type", "")
    severity = tc.get("severity", "")
    risk = tc.get("risk", "")

    # What We're Testing
    what_testing = name
    if flow_point:
        what_testing = f"{name} — verifying {flow_point}"

    # Why It Matters
    why_matters = ""
    if severity in ("S1",):
        why_matters = "Critical path; failure causes service outage or security breach"
    elif severity in ("S2",):
        why_matters = "High impact; failure degrades user experience or security posture"
    elif severity in ("S3",):
        why_matters = "Medium impact; edge case that affects specific environments"
    else:
        why_matters = "Lower priority; corner case or specific configuration"

    if related_bugs and related_bugs != "--" and "None" not in related_bugs and "Predicted" not in related_bugs:
        why_matters += f". Previously caused escalation: {related_bugs}"

    if gap_type == "Regression":
        why_matters += ". Regression risk — was working before"
    elif gap_type == "Day-1":
        why_matters += ". Day-1 gap — never tested from release"
    elif gap_type == "Test Gap":
        why_matters += ". Known test coverage gap"

    # What Could Go Wrong
    what_wrong = risk if risk else ""
    if not what_wrong:
        if "crash" in name.lower() or "crash" in flow_point.lower():
            what_wrong = "Service crash, user disruption, potential data loss"
        elif "failclose" in name.lower() or "fail close" in name.lower():
            what_wrong = "False network block, user cannot access internet"
        elif "tunnel" in name.lower():
            what_wrong = "Tunnel failure, traffic not secured"
        elif "bypass" in name.lower():
            what_wrong = "Security bypass, traffic not inspected"
        elif "cert" in name.lower() or "certificate" in name.lower():
            what_wrong = "Certificate error, HTTPS inspection failure"
        elif "upgrade" in name.lower() or "install" in name.lower():
            what_wrong = "Upgrade failure, client stuck on old version or broken state"
        elif "config" in name.lower():
            what_wrong = "Config corruption, incorrect policy applied"
        elif "npa" in name.lower():
            what_wrong = "Private app access failure"
        elif "vdi" in name.lower():
            what_wrong = "Multi-user session interference, mass user impact"
        else:
            what_wrong = f"Feature degradation in {chapter_feature}"

    return what_testing, why_matters, what_wrong


def parse_chapter(filename, sheet_name):
    """Parse a chapter and return enriched test cases."""
    filepath = os.path.join(CHAPTERS_DIR, filename)
    chapter_num = filename[:2]
    chapter_feature = get_chapter_feature(filename)

    # For chapters with pre-generated TCs, use those directly
    if chapter_num in GENERATED_TCS:
        generated = GENERATED_TCS[chapter_num]
        enriched = []
        for tc in generated:
            what_testing = tc["name"]
            why_matters = ""
            severity = tc.get("severity", "")
            if severity == "S1":
                why_matters = "Critical path; failure causes service outage or security breach"
            elif severity == "S2":
                why_matters = "High impact; failure degrades user experience or security posture"
            else:
                why_matters = "Medium impact; edge case that affects specific environments"
            if tc.get("related_bugs"):
                why_matters += f". Previously caused escalation: {tc['related_bugs']}"
            gap_type = tc.get("gap_type", "")
            if gap_type == "Regression":
                why_matters += ". Regression risk — was working before"
            elif gap_type == "Day-1":
                why_matters += ". Day-1 gap — never tested from release"
            elif gap_type == "Test Gap":
                why_matters += ". Known test coverage gap"

            row = {
                "tc_id": tc["id"],
                "name": tc["name"],
                "feature": tc.get("feature", chapter_feature),
                "what_testing": what_testing,
                "why_matters": why_matters,
                "what_wrong": tc.get("risk", ""),
                "platform": tc.get("platform", ""),
                "severity": tc.get("severity", ""),
                "auto_priority": tc.get("auto_priority", ""),
                "gap_type": tc.get("gap_type", ""),
                "related_bugs": tc.get("related_bugs", ""),
                "preconditions": tc.get("preconditions", ""),
                "steps": tc.get("steps", ""),
                "expected_result": tc.get("expected_result", ""),
                "failure_indicators": tc.get("failure_indicators", ""),
                "risk": tc.get("risk", ""),
                "automation": tc.get("automation", "❌ Not covered"),
            }
            enriched.append(row)
        return enriched

    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    # Determine format type
    detailed_chapters = {"01", "02", "10", "19", "21"}
    heading_chapters = {"04"}
    table_chapters = {"06", "08", "09", "12", "13", "14", "15", "16", "17", "18", "20"}

    test_cases = []

    if chapter_num in detailed_chapters:
        test_cases = parse_detailed_tc(content, chapter_num)
    elif chapter_num in heading_chapters:
        test_cases = parse_heading_tc(content, chapter_num)
    elif chapter_num in table_chapters:
        test_cases = parse_table_tc(content, chapter_num)

    # Also parse any cross-flow TCs (TC-CF-XX-XX) if they exist in table chapters
    if chapter_num in table_chapters:
        cf_tcs = parse_detailed_tc(content, chapter_num)
        # Only add CF TCs not already found
        existing_ids = {tc["id"] for tc in test_cases}
        for tc in cf_tcs:
            if tc["id"] not in existing_ids:
                test_cases.append(tc)

    # Enrich all test cases
    enriched = []
    for tc in test_cases:
        what_testing, why_matters, what_wrong = enrich_tc(tc, chapter_feature)

        # Fill in missing fields for table-format TCs
        steps = tc.get("steps", "")
        if not steps:
            steps = generate_steps_for_table_tc(tc, chapter_feature)

        preconditions = tc.get("preconditions", "")
        if not preconditions:
            preconditions = generate_preconditions(tc, chapter_feature)

        expected_result = tc.get("expected_result", "")
        if not expected_result:
            expected_result = generate_expected_result(tc, chapter_feature)

        failure_indicators = tc.get("failure_indicators", "")
        if not failure_indicators:
            failure_indicators = f"grep -i \"error\\|fail\\|timeout\\|crash\" nsdebuglog.log during {tc.get('name', 'test execution')}"

        automation = tc.get("automation", "")
        if not automation:
            automation = "❌ Not covered"

        row = {
            "tc_id": tc["id"],
            "name": tc.get("name", ""),
            "feature": chapter_feature,
            "what_testing": what_testing,
            "why_matters": why_matters,
            "what_wrong": what_wrong,
            "platform": tc.get("platform", ""),
            "severity": tc.get("severity", ""),
            "auto_priority": tc.get("auto_priority", ""),
            "gap_type": tc.get("gap_type", ""),
            "related_bugs": tc.get("related_bugs", ""),
            "preconditions": preconditions,
            "steps": steps,
            "expected_result": expected_result,
            "failure_indicators": failure_indicators,
            "risk": tc.get("risk", "") or what_wrong,
            "automation": automation,
        }
        enriched.append(row)

    return enriched


def create_xlsx(all_data):
    """Create xlsx workbook with styled sheets."""
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    cell_alignment = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet_name, rows in all_data:
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit 31 chars

        # Write headers
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = thin_border

        # Write data
        for row_idx, row_data in enumerate(rows, 2):
            values = [
                row_data["tc_id"],
                row_data["name"],
                row_data["feature"],
                row_data["what_testing"],
                row_data["why_matters"],
                row_data["what_wrong"],
                row_data["platform"],
                row_data["severity"],
                row_data["auto_priority"],
                row_data["gap_type"],
                row_data["related_bugs"],
                row_data["preconditions"],
                row_data["steps"],
                row_data["expected_result"],
                row_data["failure_indicators"],
                row_data["risk"],
                row_data["automation"],
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Set column widths
        for col_idx, width in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

    # Create Summary sheet at the beginning
    summary = wb.create_sheet(title="Summary", index=0)
    summary_headers = ["Sheet", "Chapter", "Test Case Count", "S1", "S2", "S3", "S4"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = summary.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border

    total = 0
    for row_idx, (sheet_name, rows) in enumerate(all_data, 2):
        s1 = sum(1 for r in rows if r["severity"] == "S1")
        s2 = sum(1 for r in rows if r["severity"] == "S2")
        s3 = sum(1 for r in rows if r["severity"] == "S3")
        s4 = sum(1 for r in rows if r["severity"] == "S4")
        count = len(rows)
        total += count

        values = [sheet_name, get_chapter_feature(sheet_name[:2] + "_"), count, s1, s2, s3, s4]
        for col_idx, val in enumerate(values, 1):
            cell = summary.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Total row
    total_row = len(all_data) + 2
    summary.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    summary.cell(row=total_row, column=3, value=total).font = Font(bold=True)

    summary.column_dimensions['A'].width = 25
    summary.column_dimensions['B'].width = 30
    summary.column_dimensions['C'].width = 15
    summary.freeze_panes = "A2"

    return wb


def main():
    print("=" * 60)
    print("NSClient Test Cases — Full Export to XLSX")
    print("=" * 60)
    print()

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

    all_data = []
    total = 0

    for filename, sheet_name in CHAPTERS:
        filepath = os.path.join(CHAPTERS_DIR, filename)
        if not os.path.exists(filepath):
            print(f"  SKIP: {filename} not found")
            continue

        rows = parse_chapter(filename, sheet_name)
        if rows:
            all_data.append((sheet_name, rows))
            total += len(rows)
            print(f"  {sheet_name}: {len(rows)} test cases")
        else:
            print(f"  {sheet_name}: 0 test cases (skipped)")

    print()
    print(f"Total: {total} test cases across {len(all_data)} sheets")
    print()

    wb = create_xlsx(all_data)
    wb.save(OUTPUT_FILE)
    print(f"Saved: {OUTPUT_FILE}")
    print()
    print("Import to Google Sheets:")
    print("  1. Open https://docs.google.com/spreadsheets/d/1ackCZ-EcepXw1BkSGoi5Go9Ex1I72-fXqcqLGMGiuio/")
    print("  2. File → Import → Upload → Select the .xlsx file")
    print("  3. Choose 'Replace spreadsheet' or 'Insert new sheets'")


if __name__ == "__main__":
    main()
